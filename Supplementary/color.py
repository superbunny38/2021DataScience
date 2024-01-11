import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def color_cells(value):
    try:
        float_value = float(value)
        return True
    except ValueError:
        return False

def read_csv_and_save_to_excel(csv_filename, excel_filename):
    # Read CSV data into a pandas DataFrame
    df = pd.read_csv(csv_filename)

    # Create a new Excel workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Write column names to the first row of the worksheet
    for c_idx, column_name in enumerate(df.columns, start=1):
        worksheet.cell(row=1, column=c_idx, value=column_name)

    # Iterate through DataFrame rows and columns
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):  # Start from row 2 to leave the first row for column names
        for c_idx, value in enumerate(row, start=1):
            # Check if the cell contains a number and apply color
            if color_cells(value):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Yellow fill for numbers
            else:
                worksheet.cell(row=r_idx, column=c_idx, value=value)  # Leave the cell as it is if not a number

    # Save the Excel workbook
    workbook.save(excel_filename)
    
import glob

csv_dir = glob.glob("./*_labeller*.csv")[0]
csv_filename = csv_dir  # Replace with your CSV file name
excel_filename = "colored_consistency_labeller.xlsx"  # Replace with your desired output Excel file name

read_csv_and_save_to_excel(csv_filename, excel_filename)